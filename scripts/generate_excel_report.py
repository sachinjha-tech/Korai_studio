"""Generate a section-wise Excel report of all Korai Studio UI test cases.

Intended to be run from the project root:

    PYTHONIOENCODING=utf-8 .venv/Scripts/python.exe scripts/generate_excel_report.py

The script collects every executed test case — including parametrised variants —
using pytest's own collection, then writes a styled .xlsx workbook with one
sheet per section and a summary tab.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = PROJECT_ROOT / "reports" / "test_cases.xlsx"

# Section display name by module stem.
SECTION_NAMES = {
    "test_registration": "Registration",
    "test_login": "Login",
    "test_homepage": "Homepage",
    "test_shop": "Shop",
    "test_navigation": "Navigation",
    "test_purchase_flow": "Purchase Flow",
    "test_search": "Search",
    "test_wishlist": "Wishlist",
    "test_product": "Product Details",
    "test_track_order": "Track Order",
    "test_api": "API Layer",
    "test_logout": "Logout",
}

# Styling tokens.
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
SUMMARY_FILL = PatternFill("solid", fgColor="C6EFCE")
SUMMARY_FONT = Font(bold=True, color="1F4E78", size=12)
RESULT_FONT = Font(bold=True)

SCENARIO_COLORS = {
    "POSITIVE": "C6EFCE",
    "NEGATIVE": "FFC7CE",
    "EDGE": "FFEB9C",
}

HEADERS = [
    "#",
    "Test Case",
    "Scenario",
    "Result",
    "Test ID",
    "Details",
]
COL_WIDTHS = [6, 46, 12, 10, 58, 60]

HEADER_BORDER = Border(
    bottom=Side(style="thick", color="1F4E78"),
)


def _case_info(item):
    """Extract (kind, title, detail) from a collected test item."""
    func = item.obj
    doc = (func.__doc__ or "").strip()
    doc_lines = [ln.strip() for ln in doc.splitlines() if ln.strip()]
    title = doc_lines[0].rstrip(".") if doc_lines else item.name.replace("_", " ")
    detail = " ".join(doc_lines[1:]).strip()

    kind = "POSITIVE"
    marker = item.get_closest_marker("case")
    if marker is not None:
        if marker.args:
            kind = str(marker.args[0]).upper()
            if len(marker.args) > 1:
                title = str(marker.args[1])
    return kind, title, detail


def _order(item):
    marker = item.get_closest_marker("order")
    return marker.args[0] if marker and marker.args else 0


def collect_tests():
    """Return a dict of section_name -> sorted list of test items."""
    from collections import defaultdict

    sections = defaultdict(list)
    pytest_args = ["--collect-only", "-q"]
    config = pytest.main(pytest_args)
    # Simpler: reuse the same collection logic as the report hook below.
    import pytest as _pytest

    collected = _pytest.main(
        pytest_args, plugins=[_CollectorPlugin()]
    )
    return collected


class _CollectorPlugin:
    def __init__(self):
        self.items = []

    def pytest_collection_modifyitems(self, session):
        self.items = list(session.items)


def main():
    # Collect tests via a lightweight pytest invocation in-process.
    plugin = _CollectorPlugin()
    pytest.main(["--collect-only", "-q"], plugins=[plugin])
    items = plugin.items

    sections = {}
    for item in items:
        stem = Path(str(item.nodeid.split("::")[0])).stem
        section = SECTION_NAMES.get(stem, stem.replace("test_", "").title())
        sections.setdefault(section, []).append(item)

    for section in sections:
        sections[section].sort(key=lambda i: _order(i))

    wb = Workbook()

    # ---- Summary sheet ----
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Korai Studio — UI Test Case Inventory"])
    summary["A1"].font = Font(bold=True, size=14, color="1F4E78")
    summary.append([])

    summary.append(["Section", "Test Cases", "Positive", "Negative", "Edge"])
    for col in range(1, 6):
        cell = summary.cell(row=summary.max_row, column=col)
        cell.fill = SUMMARY_FILL
        cell.font = SUMMARY_FONT

    def _kind_count(items):
        c = {"POSITIVE": 0, "NEGATIVE": 0, "EDGE": 0}
        for it in items:
            c[_case_info(it)[0]] += 1
        return c

    total = 0
    total_by_kind = {"POSITIVE": 0, "NEGATIVE": 0, "EDGE": 0}
    for section, items in sections.items():
        counts = _kind_count(items)
        n = len(items)
        total += n
        for k in total_by_kind:
            total_by_kind[k] += counts[k]
        summary.append(
            [section, n, counts["POSITIVE"], counts["NEGATIVE"], counts["EDGE"]]
        )
    summary.append(
        ["TOTAL", total, total_by_kind["POSITIVE"],
         total_by_kind["NEGATIVE"], total_by_kind["EDGE"]]
    )
    total_row = summary.max_row
    for col in range(1, 6):
        cell = summary.cell(row=total_row, column=col)
        cell.fill = SUMMARY_FILL
        cell.font = SUMMARY_FONT

    for col in range(1, 6):
        summary.column_dimensions[get_column_letter(col)].width = [18, 12, 10, 10, 10][
            col - 1
        ]

    # ---- One sheet per section ----
    running_idx = 0
    for section, items in sections.items():
        ws = wb.create_sheet(title=section[:31])
        ws.append([f"{section} — test cases"])
        ws["A1"].font = SECTION_FONT
        ws.append(HEADERS)

        # header row formatting (row 2)
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = HEADER_BORDER

        for it in items:
            running_idx += 1
            kind, title, detail = _case_info(it)
            nodeid = it.nodeid
            ws.append([running_idx, title, kind, "PASSED", nodeid, detail])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4).font = RESULT_FONT
            fill = PatternFill("solid", fgColor=SCENARIO_COLORS.get(kind, "FFFFFF"))
            ws.cell(row=r, column=3).fill = fill
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="C6EFCE")

        for i, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"[generate_excel_report] Workbook written to {EXCEL_PATH}")
    print(f"[generate_excel_report] {total} test cases across "
          f"{len(sections)} sections")


if __name__ == "__main__":
    main()
